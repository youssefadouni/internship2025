import os
import logging
import pytesseract
import pdfplumber
from docx import Document
from openpyxl import load_workbook
from PIL import Image  # Required for handling images during OCR

# Set up logging to both file and console for debugging and audit trail
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler("extraction.log"),
                        logging.StreamHandler()
                    ])

UPLOADS_DIR = "uploads"

def extract_text_from_pdf(file_path):
    """Extract text from PDF files with OCR fallback for scanned documents.
    
    First attempts direct text extraction, then falls back to OCR if no text found.
    OCR is more resource-intensive but handles scanned documents.
    """
    text = ""
    try:
        with pdfplumber.open(file_path) as pdf:
            # First pass: Try native text extraction (faster, works for digital PDFs)
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            
            # Second pass: If no text found, PDF might be scanned/image-based
            # Fall back to OCR which is slower but can handle image-based PDFs
            if not text.strip():
                logging.info(f"No text directly extracted from {file_path}. Attempting OCR.")
                # TODO: Consider setting tesseract path in a config file instead of hardcoding
                # pytesseract.pytesseract.tesseract_cmd = r'/usr/local/bin/tesseract'
                full_ocr_text = ""
                for i, page in enumerate(pdf.pages):
                    try:
                        # Convert page to image at 300dpi - tradeoff between quality and performance
                        # Higher resolution = better OCR but slower processing and more memory usage
                        im = page.to_image(resolution=300)
                        
                        # Default to English - in production, consider language detection
                        # or allowing language specification via parameters
                        page_ocr_text = pytesseract.image_to_string(im.original, lang='eng')
                        if page_ocr_text:
                            full_ocr_text += page_ocr_text + "\n"
                        logging.info(f"OCR processed page {i+1} of {file_path}")
                    except Exception as ocr_page_error:
                        # Continue processing other pages even if one fails
                        logging.error(f"Error during OCR for page {i+1} of {file_path}: {ocr_page_error}")
                text = full_ocr_text
                if text.strip():
                    logging.info(f"Successfully extracted text using OCR from {file_path}")
                else:
                    logging.warning(f"OCR attempt on {file_path} yielded no text.")

    except Exception as e:
        logging.error(f"Error processing PDF file {file_path}: {e}")
    return text.strip()

def extract_text_from_docx(file_path):
    """Extract text from Word documents (.docx).
    
    Extracts text from paragraphs only. Does not handle tables, headers, footers,
    or other special elements - extend if needed for your specific documents.
    """
    text = ""
    try:
        doc = Document(file_path)
        for para in doc.paragraphs:
            text += para.text + "\n"
        # TODO: I will be adding support for tables, headers, footers if needed
    except Exception as e:
        logging.error(f"Error processing DOCX file {file_path}: {e}")
    return text.strip()

def extract_text_from_xlsx(file_path):
    """Extract text from Excel spreadsheets (.xlsx).
    
    Uses read_only and data_only modes for better performance with large files.
    Preserves row structure but not cell formatting or formulas.
    """
    text = ""
    try:
        # read_only=True improves performance for large files
        # data_only=True gets calculated values instead of formulas
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        text += str(cell.value) + " "
                text += "\n"
    except Exception as e:
        logging.error(f"Error processing XLSX file {file_path}: {e}")
    return text.strip()

def process_files_in_uploads():
    """Process all supported files in the uploads directory.
    
    Supports PDF, DOCX, and XLSX files. Returns a dictionary mapping
    filenames to their extracted text content.
    """
    if not os.path.exists(UPLOADS_DIR):
        logging.error(f"Directory not found: {UPLOADS_DIR}")
        print(f"Error: Directory '{UPLOADS_DIR}' not found. Please create it and add files.")
        return
    
    if not os.listdir(UPLOADS_DIR):
        logging.info(f"No files found in {UPLOADS_DIR} directory.")
        print(f"No files found in '{UPLOADS_DIR}'. Add some .pdf, .docx, or .xlsx files to process.")
        return

    extracted_data = {}
    for filename in os.listdir(UPLOADS_DIR):
        file_path = os.path.join(UPLOADS_DIR, filename)
        if os.path.isfile(file_path):
            logging.info(f"Processing file: {filename}")
            extracted_text = ""
            
            # Dispatch to appropriate handler based on file extension
            # Could be refactored to use a strategy pattern for more file types
            if filename.lower().endswith(".pdf"):
                extracted_text = extract_text_from_pdf(file_path)
            elif filename.lower().endswith(".docx"):
                extracted_text = extract_text_from_docx(file_path)
            elif filename.lower().endswith(".xlsx"):
                extracted_text = extract_text_from_xlsx(file_path)
            else:
                logging.info(f"Skipping unsupported file type: {filename}")
                continue
            
            if extracted_text:
                extracted_data[filename] = extracted_text
                # Uncomment for debugging or to see full text output
                # print(f"--- Text from {filename} ---\n{extracted_text}\n---------------------------\n")
            else:
                logging.warning(f"No text extracted from {filename}")
        else:
            logging.info(f"Skipping non-file item: {filename}")
    
    # In a real pipeline, we might want to save to a database or pass to NLP processing
    if extracted_data:
        logging.info("Text extraction complete for all supported files.")
        # Example of how we might save all text to a single file:
        # with open("all_extracted_text.txt", "w", encoding="utf-8") as f:
        #     for filename, text in extracted_data.items():
        #         f.write(f"--- {filename} ---\n{text}\n\n")
    else:
        logging.info("No text was extracted from any files.")

    return extracted_data

if __name__ == "__main__":
    # Tesseract dependency check - this is a common source of issues
    # Installation instructions for reference:
    # macOS: brew install tesseract
    # Ubuntu: apt-get install tesseract-ocr
    # Windows: https://github.com/UB-Mannheim/tesseract/wiki
    
    try:
        tesseract_version = pytesseract.get_tesseract_version()
        logging.info(f"Tesseract version {tesseract_version} found.")
    except pytesseract.TesseractNotFoundError:
        logging.error("Tesseract is not installed or not found in your PATH.")
        logging.error("Please install Tesseract and ensure it's in your system's PATH or set pytesseract.tesseract_cmd.")
        # Uncomment to force exit if Tesseract is required for your workflow
        # exit(1) 

    logging.info("Starting text extraction process...")
    all_texts = process_files_in_uploads()
    
    # Print a summary of extracted text for user feedback
    if all_texts:
        print("\n--- Summary of Extracted Text ---")
        for filename, text_preview in all_texts.items():
            print(f"File: {filename}")
            # Show preview of first 200 chars to avoid flooding the console
            print(f"Preview: {text_preview[:200]}..." if len(text_preview) > 200 else text_preview)
            print("-----------------------------------")
    else:
        print("No text could be extracted from the files in the 'uploads' directory.")
    
    logging.info("Text extraction process finished.")